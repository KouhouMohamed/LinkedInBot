import fonctions
import requests
from bs4 import BeautifulSoup
import openpyxl
from secrets import username, password
import sqlite3
        

if __name__ == '__main__':
    # Enmplacement de fichier xlsx ou le résultat est stocker 
    
    path_xlsx = r".\ListOfStudents.xlsx"
    work_book = openpyxl.load_workbook(path_xlsx)
    sheet = work_book.active
    row = 1
    colum = 1
    # first row in the table
    sheet.cell(row=row, column = colum).value = "Student"
    sheet.cell(row=row, column = colum + 1).value = "School"
    sheet.cell(row=row, column = colum + 2).value = "Image"
    
    ListOfStudents =['Enter names of students you are looking for there schools']
    #create an objects from the class LinkedinBot 
    bot = fonctions.LinkedinBot(username, password)

    #Login to Linkedin with the inputs informations
    bot.login()
    
    for student in ListOfStudents:
        
        # go to the next row to fill it
        row = row + 1
        liste = bot.search(student)
        if len(liste) == 0:
            bot.school = 'account not found'
            bot.image = 'account not found'
            bot._nav(bot.feed_url)
        else:
            url = liste[0]
            bot.searchSchool(url)
        # fill the row with the name of student and his school name
        sheet.cell(row=row, column = colum).value = student
        sheet.cell(row=row, column = colum + 1).value = bot.school
        sheet.cell(row=row, column = colum + 2).value = bot.image
        
        #save the work
        work_book.save(path_xlsx)
    bot.quit()